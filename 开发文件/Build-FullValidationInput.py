from __future__ import annotations

import argparse
import csv
from pathlib import Path
from urllib.parse import urlsplit

from openpyxl import load_workbook


OUTPUT_HEADERS = [
    "序号", "平台名称", "标题", "摘要", "链接", "账号昵称", "内容类型",
    "历史处置", "来源工作表", "来源行号", "数据集分组",
]


def text(value) -> str:
    return "" if value is None else str(value).strip()


def normalized(value) -> str:
    return text(value).replace(" ", "").lower()


def find_column(headers, candidates):
    wanted = {normalized(value) for value in candidates}
    return next((index for index, value in enumerate(headers) if normalized(value) in wanted), -1)


def value_at(row, index):
    return row[index] if 0 <= index < len(row) else None


def is_http_url(value) -> bool:
    try:
        parsed = urlsplit(text(value))
    except ValueError:
        return False
    return parsed.scheme.lower() in {"http", "https"} and bool(parsed.hostname)


def main():
    parser = argparse.ArgumentParser(description="Build a complete local validation CSV from an XLSX workbook.")
    parser.add_argument("input_xlsx", type=Path)
    parser.add_argument("output_csv", type=Path)
    parser.add_argument("--expected-rows", type=int, default=0)
    parser.add_argument("--original-sheet", default="")
    parser.add_argument("--expected-original-rows", type=int, default=0)
    args = parser.parse_args()

    input_path = args.input_xlsx.resolve()
    output_path = args.output_csv.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    total = 0
    original_rows = 0
    sheet_counts = {}

    with output_path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=OUTPUT_HEADERS)
        writer.writeheader()
        for worksheet in workbook.worksheets:
            worksheet.reset_dimensions()
            rows = worksheet.iter_rows(values_only=True)
            headers = list(next(rows, ()))
            link_column = find_column(headers, ("链接", "网址", "原链接", "URL"))
            if link_column < 0:
                continue
            platform_column = find_column(headers, ("发布平台", "平台名称", "来源平台", "平台"))
            title_column = find_column(headers, ("标题", "文章标题", "内容标题"))
            excerpt_column = find_column(headers, ("摘要", "内容摘要", "正文摘要"))
            author_column = find_column(headers, ("发布人", "账号昵称", "作者", "发布账号"))
            status_column = find_column(headers, ("是否删除", "核验是否下架", "处置情况", "链接是否失效"))
            count = 0
            for source_row, row in enumerate(rows, 2):
                values = list(row)
                url = text(value_at(values, link_column))
                if not is_http_url(url):
                    continue
                total += 1
                count += 1
                if worksheet.title == args.original_sheet:
                    original_rows += 1
                writer.writerow({
                    "序号": total,
                    "平台名称": text(value_at(values, platform_column)),
                    "标题": text(value_at(values, title_column)),
                    "摘要": text(value_at(values, excerpt_column)),
                    "链接": url,
                    "账号昵称": text(value_at(values, author_column)),
                    "内容类型": "",
                    "历史处置": text(value_at(values, status_column)),
                    "来源工作表": worksheet.title,
                    "来源行号": source_row,
                    "数据集分组": "原始附件" if worksheet.title == args.original_sheet else "后续新增",
                })
            sheet_counts[worksheet.title] = count
    workbook.close()

    if args.expected_rows and total != args.expected_rows:
        raise SystemExit(f"row count mismatch: {total} != {args.expected_rows}")
    if args.expected_original_rows and original_rows != args.expected_original_rows:
        raise SystemExit(f"original row count mismatch: {original_rows} != {args.expected_original_rows}")
    print(f"FULL_VALIDATION_ROWS={total}")
    print(f"ORIGINAL_ATTACHMENT_ROWS={original_rows}")
    for sheet, count in sheet_counts.items():
        print(f"SHEET={sheet},ROWS={count}")
    print(f"OUTPUT={output_path}")


if __name__ == "__main__":
    main()
