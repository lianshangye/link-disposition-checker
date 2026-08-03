from __future__ import annotations

import argparse
import csv
import hashlib
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import parse_qsl, urlencode, urlsplit

from openpyxl import load_workbook


OUTPUT_HEADERS = [
    "序号", "平台名称", "标题", "摘要", "账号昵称", "链接", "内容类型", "处置情况",
    "来源工作表", "来源行号", "标签快照时间", "发文时间", "审计种子",
]
HISTORY_HEADERS = OUTPUT_HEADERS + ["抽样时间"]
URL_COLUMNS = ("链接", "网址", "原链接", "url", "URL")
TRACKING_PARAMETERS = {
    "spm", "scm", "refer_flag", "share_token", "share_source", "share_from", "isappinstalled",
}


def text(value):
    return "" if value is None else str(value).strip()


def value_at(row, index):
    return row[index] if index < len(row) else None


def canonical_url(value):
    try:
        parsed = urlsplit(text(value))
    except ValueError:
        return ""
    if parsed.scheme.lower() not in ("http", "https") or not parsed.hostname:
        return ""
    host = parsed.hostname.lower()
    if parsed.port and not ((parsed.scheme.lower() == "http" and parsed.port == 80) or
                            (parsed.scheme.lower() == "https" and parsed.port == 443)):
        host += ":" + str(parsed.port)
    path = parsed.path or "/"
    if len(path) > 1:
        path = path.rstrip("/")
    query = []
    for key, value in parse_qsl(parsed.query, keep_blank_values=True):
        normalized = key.strip().lower()
        if normalized.startswith("utm_") or normalized in TRACKING_PARAMETERS:
            continue
        query.append((key, value))
    query.sort(key=lambda pair: (pair[0], pair[1]))
    return host + path + (("?" + urlencode(query)) if query else "")


def score(seed, *values):
    payload = "|".join((seed,) + tuple(text(value) for value in values))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def read_excluded_urls(paths):
    excluded = set()
    for path in paths:
        if not path.exists() or path.stat().st_size == 0:
            continue
        with path.open("r", encoding="utf-8-sig", newline="", errors="replace") as stream:
            reader = csv.DictReader(stream)
            if not reader.fieldnames:
                continue
            url_column = next((name for name in URL_COLUMNS if name in reader.fieldnames), None)
            if not url_column:
                continue
            for row in reader:
                key = canonical_url(row.get(url_column, ""))
                if key:
                    excluded.add(key)
    return excluded


def read_used_seeds(history_path):
    if not history_path.exists() or history_path.stat().st_size == 0:
        return set()
    with history_path.open("r", encoding="utf-8-sig", newline="", errors="replace") as stream:
        return {text(row.get("审计种子")) for row in csv.DictReader(stream) if text(row.get("审计种子"))}


def load_records(input_path):
    received_at = datetime.fromtimestamp(input_path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    configs = [
        (0, 6, 7, 8, 10, 11, 12, 3, "2026-07-18 20:00:00"),
        (1, 7, 8, 9, 10, 11, 12, 4, received_at),
    ]
    records = []
    for sheet_index, platform_index, title_index, summary_index, url_index, status_index, date_index, author_index, snapshot_at in configs:
        worksheet = workbook.worksheets[sheet_index]
        worksheet.reset_dimensions()
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), 1):
            if row_number == 1:
                continue
            values = list(row)
            status = text(value_at(values, status_index))
            if status not in ("下架", "否"):
                continue
            url = text(value_at(values, url_index))
            key = canonical_url(url)
            if not key:
                continue
            published = value_at(values, date_index)
            records.append({
                "platform": text(value_at(values, platform_index)) or "未知平台",
                "title": text(value_at(values, title_index)),
                "summary": text(value_at(values, summary_index)),
                "author": text(value_at(values, author_index)),
                "url": url,
                "url_key": key,
                "status": status,
                "source_sheet": worksheet.title,
                "source_row": row_number,
                "snapshot_at": snapshot_at,
                "published_at": published.strftime("%Y-%m-%d %H:%M:%S") if isinstance(published, datetime) else "",
            })
    workbook.close()
    return records


def remove_ambiguous_and_duplicate_records(records, seed):
    by_url = defaultdict(list)
    for record in records:
        by_url[record["url_key"]].append(record)
    unique = []
    ambiguous = 0
    for group in by_url.values():
        if len({record["status"] for record in group}) != 1:
            ambiguous += 1
            continue
        unique.append(min(group, key=lambda record: score(seed, record["source_sheet"], record["source_row"])))
    return unique, ambiguous


def select_balanced(records, maximum_rows, seed):
    status_targets = {
        "下架": maximum_rows // 2,
        "否": maximum_rows - (maximum_rows // 2),
    }
    selected = []
    for status in ("下架", "否"):
        groups = defaultdict(list)
        for record in records:
            if record["status"] == status:
                groups[record["platform"]].append(record)
        queues = []
        for platform, group in groups.items():
            ordered = sorted(group, key=lambda record: score(seed, status, platform, record["url_key"]))
            queues.append((score(seed, status, platform), ordered))
        queues.sort(key=lambda item: item[0])
        target = status_targets[status]
        while len([row for row in selected if row["status"] == status]) < target:
            progress = False
            for _, queue in queues:
                if not queue:
                    continue
                selected.append(queue.pop(0))
                progress = True
                if len([row for row in selected if row["status"] == status]) >= target:
                    break
            if not progress:
                break
    return sorted(selected, key=lambda record: score(seed, record["status"], record["platform"], record["url_key"]))


def output_row(number, record, seed, sampled_at=None):
    row = {
        "序号": number,
        "平台名称": record["platform"],
        "标题": record["title"],
        "摘要": record["summary"],
        "账号昵称": record["author"],
        "链接": record["url"],
        "内容类型": "",
        "处置情况": record["status"],
        "来源工作表": record["source_sheet"],
        "来源行号": record["source_row"],
        "标签快照时间": record["snapshot_at"],
        "发文时间": record["published_at"],
        "审计种子": seed,
    }
    if sampled_at is not None:
        row["抽样时间"] = sampled_at
    return row


def write_output(path, selected, seed):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=OUTPUT_HEADERS)
        writer.writeheader()
        for number, record in enumerate(selected, 1):
            writer.writerow(output_row(number, record, seed))


def append_history(path, selected, seed):
    path.parent.mkdir(parents=True, exist_ok=True)
    needs_header = not path.exists() or path.stat().st_size == 0
    sampled_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    existing_rows = 0
    if not needs_header:
        with path.open("r", encoding="utf-8-sig", newline="", errors="replace") as stream:
            existing_rows = sum(1 for _ in csv.DictReader(stream))
    with path.open("a", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=HISTORY_HEADERS)
        if needs_header:
            writer.writeheader()
        for offset, record in enumerate(selected, 1):
            writer.writerow(output_row(existing_rows + offset, record, seed, sampled_at))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("input_xlsx", type=Path)
    parser.add_argument("output_csv", type=Path)
    parser.add_argument("history_csv", type=Path)
    parser.add_argument("--seed", required=True)
    parser.add_argument("--rows", type=int, default=300)
    parser.add_argument("--exclude", action="append", default=[], type=Path)
    args = parser.parse_args()

    input_path = args.input_xlsx.resolve()
    output_path = args.output_csv.resolve()
    history_path = args.history_csv.resolve()
    if args.rows < 50:
        raise SystemExit("accuracy audit requires at least 50 rows")
    used_seeds = read_used_seeds(history_path)
    if args.seed in used_seeds:
        raise SystemExit("accuracy audit seed has already been used: " + args.seed)

    exclude_paths = [history_path] + [path.resolve() for path in args.exclude]
    excluded = read_excluded_urls(exclude_paths)
    records, ambiguous = remove_ambiguous_and_duplicate_records(load_records(input_path), args.seed)
    available = [record for record in records if record["url_key"] not in excluded]
    selected = select_balanced(available, args.rows, args.seed)
    if len(selected) != args.rows:
        raise SystemExit("insufficient unused balanced human-label rows: " + str(len(selected)) + " < " + str(args.rows))
    if len({record["url_key"] for record in selected}) != len(selected):
        raise SystemExit("accuracy audit selected duplicate normalized URLs")
    if any(record["url_key"] in excluded for record in selected):
        raise SystemExit("accuracy audit reused an excluded normalized URL")

    write_output(output_path, selected, args.seed)
    append_history(history_path, selected, args.seed)
    print("ACCURACY_SOURCE_LABELLED=" + str(len(records)))
    print("ACCURACY_AMBIGUOUS_URLS=" + str(ambiguous))
    print("ACCURACY_EXCLUDED_URLS=" + str(len(excluded)))
    print("ACCURACY_AVAILABLE_URLS=" + str(len(available)))
    print("ACCURACY_SAMPLE_ROWS=" + str(len(selected)))
    print("ACCURACY_SAMPLE_REMOVED=" + str(sum(1 for row in selected if row["status"] == "下架")))
    print("ACCURACY_SAMPLE_ALIVE=" + str(sum(1 for row in selected if row["status"] == "否")))
    print("ACCURACY_SAMPLE_PLATFORMS=" + str(len({row["platform"] for row in selected})))
    print("ACCURACY_REUSED_URLS=0")
    print("ACCURACY_OUTPUT=" + str(output_path))
    print("ACCURACY_HISTORY=" + str(history_path))


if __name__ == "__main__":
    main()
