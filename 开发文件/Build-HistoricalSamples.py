from __future__ import annotations

import csv
import hashlib
import math
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def text(value):
    return "" if value is None else str(value).strip()


def value_at(row, index):
    return row[index] if index < len(row) else None


def record_key(record):
    published = record["published_at"]
    return (published if isinstance(published, datetime) else datetime.min, record["source_row"])


def write_csv(path, records):
    headers = [
        "序号", "平台名称", "标题", "摘要", "链接", "账号昵称", "处置情况",
        "来源工作表", "来源行号", "标签快照时间", "发文时间",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=headers)
        writer.writeheader()
        for number, record in enumerate(records, 1):
            writer.writerow({
                "序号": number,
                "平台名称": record["platform"],
                "标题": record["title"],
                "摘要": record["summary"],
                "链接": record["url"],
                "账号昵称": record["author"],
                "处置情况": record["historical_status"],
                "来源工作表": record["source_sheet"],
                "来源行号": record["source_row"],
                "标签快照时间": record["snapshot_at"],
                "发文时间": record["published_at"].strftime("%Y-%m-%d %H:%M:%S") if isinstance(record["published_at"], datetime) else "",
            })


def stable_score(record):
    value = "|".join((
        record["platform"], record["historical_status"], record["url"],
        str(record["source_sheet"]), str(record["source_row"]),
    ))
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def build_stratified_sample(records, target_size=600):
    """Sample every platform/status stratum, weighted toward larger platforms."""
    groups = defaultdict(list)
    for record in records:
        groups[(record["platform"], record["historical_status"])].append(record)

    quotas = {}
    for key, group in groups.items():
        # Small strata keep two examples; large strata grow sub-linearly so a few
        # dominant platforms do not crowd all long-tail platforms out of the audit.
        quotas[key] = min(len(group), max(2, int(math.ceil(math.sqrt(len(group))))))

    while sum(quotas.values()) > target_size:
        reducible = [key for key in quotas if quotas[key] > 2]
        if not reducible:
            break
        key = max(reducible, key=lambda item: quotas[item])
        quotas[key] -= 1
    while sum(quotas.values()) < target_size:
        expandable = [key for key, group in groups.items() if quotas[key] < len(group)]
        if not expandable:
            break
        key = max(expandable, key=lambda item: len(groups[item]) / float(quotas[item] + 1))
        quotas[key] += 1

    selected = []
    for key, group in groups.items():
        ordered = sorted(group, key=stable_score)
        selected.extend(ordered[:quotas[key]])
    return sorted(selected, key=lambda item: (item["platform"], item["historical_status"], stable_score(item)))


def build_representative_sample(records, target_size=1200):
    """Stable uniform sample that preserves the workbook's real platform mix."""
    if len(records) <= target_size:
        return sorted(records, key=stable_score)
    return sorted(records, key=stable_score)[:target_size]


def main():
    if len(sys.argv) < 3:
        raise SystemExit("usage: Build-HistoricalSamples.py input.xlsx output-directory")
    input_path = Path(sys.argv[1]).resolve()
    output_directory = Path(sys.argv[2]).resolve()
    output_directory.mkdir(parents=True, exist_ok=True)
    received_at = datetime.fromtimestamp(input_path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    configs = [
        (0, 6, 7, 8, 10, 11, 12, 3, "2026-07-18 20:00:00"),
        (1, 7, 8, 9, 10, 11, 12, 4, received_at),
    ]
    records = []
    status_counts = Counter()
    sheet_counts = Counter()
    platform_counts = Counter()

    for sheet_index, platform_index, title_index, summary_index, url_index, status_index, date_index, author_index, snapshot_at in configs:
        worksheet = workbook.worksheets[sheet_index]
        worksheet.reset_dimensions()
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), 1):
            if row_number == 1:
                continue
            values = list(row)
            url = text(value_at(values, url_index))
            if not url:
                continue
            status = text(value_at(values, status_index))
            platform = text(value_at(values, platform_index))
            status_counts[status] += 1
            sheet_counts[worksheet.title] += 1
            platform_counts[platform] += 1
            if status not in ("下架", "否"):
                continue
            records.append({
                "platform": platform,
                "title": text(value_at(values, title_index)),
                "summary": text(value_at(values, summary_index)),
                "url": url,
                "author": text(value_at(values, author_index)),
                "historical_status": status,
                "source_sheet": worksheet.title,
                "source_row": row_number,
                "snapshot_at": snapshot_at,
                "published_at": value_at(values, date_index),
                "sheet_index": sheet_index,
            })

    latest_by_platform_status = {}
    for record in records:
        key = (record["platform"], record["historical_status"])
        if key not in latest_by_platform_status or record_key(record) > record_key(latest_by_platform_status[key]):
            latest_by_platform_status[key] = record
    historical_samples = sorted(latest_by_platform_status.values(), key=lambda item: (item["platform"], item["historical_status"]))

    latest_recent_by_platform = {}
    for record in records:
        if record["sheet_index"] != 1:
            continue
        key = record["platform"]
        if key not in latest_recent_by_platform or record_key(record) > record_key(latest_recent_by_platform[key]):
            latest_recent_by_platform[key] = record
    recent_samples = sorted(latest_recent_by_platform.values(), key=lambda item: item["platform"])
    stratified_samples = build_stratified_sample(records)
    representative_samples = build_representative_sample(records)

    historical_path = output_directory / "historical-human-samples.csv"
    recent_path = output_directory / "recent-platform-samples.csv"
    stratified_path = output_directory / "stratified-validation-samples.csv"
    representative_path = output_directory / "representative-validation-samples.csv"
    write_csv(historical_path, historical_samples)
    write_csv(recent_path, recent_samples)
    write_csv(stratified_path, stratified_samples)
    write_csv(representative_path, representative_samples)

    summary_lines = [
        "Historical human-label sample inventory",
        "Source: " + str(input_path),
        "Workbook received/modified: " + received_at,
        "Total URL rows: " + str(sum(sheet_counts.values())),
        "Usable historical labels (down/no): " + str(len(records)),
        "Platform/status samples: " + str(len(historical_samples)),
        "Recent one-per-platform samples: " + str(len(recent_samples)),
        "Stratified validation samples: " + str(len(stratified_samples)),
        "Representative validation samples: " + str(len(representative_samples)),
        "",
        "Important: labels are historical observations, not permanent truth.",
        "The workbook has no per-row verification timestamp. Sheet 1 uses its named cutoff; sheet 2 uses the workbook received time.",
        "Statuses other than down/no are excluded from link-state ground truth.",
        "",
        "Status counts:",
    ]
    for status, count in status_counts.most_common():
        summary_lines.append("- " + (status or "(blank)") + ": " + str(count))
    summary_lines.extend(["", "Largest platforms:"])
    for platform, count in platform_counts.most_common(30):
        summary_lines.append("- " + platform + ": " + str(count))
    (output_directory / "historical-human-summary.txt").write_text("\n".join(summary_lines) + "\n", encoding="utf-8-sig")
    print("historical samples", len(historical_samples), historical_path)
    print("recent samples", len(recent_samples), recent_path)
    print("stratified samples", len(stratified_samples), stratified_path)
    print("representative samples", len(representative_samples), representative_path)


if __name__ == "__main__":
    main()
